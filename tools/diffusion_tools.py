import io
import json
import os
import requests
from crewai.tools import tool
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

SCOPES = ["https://www.googleapis.com/auth/drive"]
BREVO_API_URL = "https://api.brevo.com/v3/smtp/email"

FOLDER_DIFFUSION = "1peQ9728pAY2h2j60i-Wns4-bXyjZzVzl"


def _get_drive_service():
    creds = service_account.Credentials.from_service_account_file(
        os.environ["GOOGLE_SERVICE_ACCOUNT_FILE"], scopes=SCOPES
    )
    return build("drive", "v3", credentials=creds)


def _download_bytes(service, file_id: str) -> bytes:
    request = service.files().get_media(fileId=file_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buffer.getvalue()


def _extraire_objet_et_corps(texte: str) -> tuple[str, str]:
    """
    Extrait l'objet et le corps depuis le texte brut de ContenuMessage.
    Convention attendue :
        Objet : [ligne contenant l'objet de l'email]  ← espace avant et après ':'
        [ligne vide optionnelle]
        [corps du message]

    Accepte aussi 'Objet:' sans espace avant le ':'.
    Si aucune ligne "Objet" n'est trouvée, la première ligne devient l'objet.
    """
    import re
    lines = texte.splitlines()
    objet = ""
    corps_lines = []
    objet_trouve = False

    # Détecte "Objet :" ou "Objet:" en début de ligne (insensible à la casse)
    _OBJET_RE = re.compile(r"^objet\s*:", re.IGNORECASE)

    for i, line in enumerate(lines):
        if not objet_trouve and _OBJET_RE.match(line.strip()):
            # Extrait ce qui suit le premier ':'
            objet = line.split(":", 1)[1].strip()
            objet_trouve = True
            corps_lines = [ln for ln in lines[i + 1:] if ln.strip() or corps_lines]
            break

    if not objet_trouve:
        objet = lines[0].strip() if lines else ""
        corps_lines = lines[1:]

    corps = "\n".join(corps_lines).strip()
    return objet, corps


@tool("Lire le fichier ContenuMessage depuis Google Drive")
def lire_contenu_message_depuis_drive() -> str:
    """
    Recherche dynamiquement le fichier ContenuMessage (tout format : .docx, .txt, .pdf)
    dans le dossier Drive Diffusion_et_Communication par son nom, puis l'analyse.
    Retourne un JSON {"objet": "...", "corps": "..."} selon la convention :
      - La ligne commençant par 'Objet :' contient le sujet de l'email.
      - Le reste du fichier constitue le corps du message.
    """
    service = _get_drive_service()

    results = service.files().list(
        q=(
            f"'{FOLDER_DIFFUSION}' in parents "
            "and name contains 'ContenuMessage' "
            "and trashed=false"
        ),
        fields="files(id, name)",
    ).execute()
    files = results.get("files", [])
    if not files:
        return json.dumps(
            {"erreur": "Fichier ContenuMessage introuvable dans Diffusion_et_Communication."})

    file_id = files[0]["id"]
    name = files[0]["name"].lower()
    data = _download_bytes(service, file_id)

    if name.endswith(".txt"):
        texte = data.decode("utf-8", errors="ignore").strip()

    elif name.endswith(".docx"):
        from docx import Document
        doc = Document(io.BytesIO(data))
        texte = "\n".join(p.text for p in doc.paragraphs)

    elif name.endswith(".pdf"):
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            texte = "\n".join(page.extract_text() or "" for page in pdf.pages).strip()

    else:
        return json.dumps({
            "erreur": f"Format non supporté ({files[0]['name']}). Utiliser .txt, .docx ou .pdf."
        })

    objet, corps = _extraire_objet_et_corps(texte)
    return json.dumps({"objet": objet, "corps": corps}, ensure_ascii=False)


@tool("Lire les emails depuis le fichier Excel ListeContacts_Lin_Out_FINAL")
def lire_emails_depuis_excel_drive(folder_id: str) -> str:
    """
    Lit le fichier Excel 'ListeContacts_Lin_Out_FINAL_jjmmaaaa.xlsx' dans le dossier
    Google Drive spécifié (recherche par préfixe 'ListeContacts_Lin_Out_FINAL').
    Retourne un JSON {"fichier": "nom_exact.xlsx", "emails": [...]}
    où 'fichier' est le nom exact trouvé (utile pour l'archivage) et
    'emails' est la liste des adresses valides extraites de la colonne 'Email'.
    """
    service = _get_drive_service()

    results = service.files().list(
        q=(
            f"'{folder_id}' in parents "
            "and name contains 'ListeContacts_Lin_Out_FINAL' "
            "and trashed=false"
        ),
        fields="files(id, name)",
    ).execute()
    files = results.get("files", [])
    if not files:
        return json.dumps({"erreur": "Fichier ListeContacts_Lin_Out_FINAL introuvable."})

    fichier_nom = files[0]["name"]
    data = _download_bytes(service, files[0]["id"])

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    try:
        col_email = headers.index("Email")
    except ValueError:
        return json.dumps({"erreur": "Colonne 'Email' introuvable dans le fichier Excel."})

    emails = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_email] if len(row) > col_email else None
        if val and "@" in str(val):
            emails.append(str(val).strip().lower())

    return json.dumps({"fichier": fichier_nom, "emails": emails}, ensure_ascii=False)


@tool("Envoyer un message identique à une liste d'emails via Sendinblue (Brevo)")
def envoyer_emails_via_sendinblue(emails_json: str, sujet: str, contenu: str) -> str:
    """
    Envoie un email identique à chaque adresse de emails_json via l'API Brevo (Sendinblue).
    - emails_json : JSON array de chaînes email
    - sujet       : objet de l'email (extrait de ContenuMessage)
    - contenu     : corps du message en texte brut (extrait de ContenuMessage)

    Variables d'environnement requises :
      SENDINBLUE_API_KEY      — clé API Brevo
      SENDINBLUE_SENDER_EMAIL — adresse expéditeur
      SENDINBLUE_SENDER_NAME  — nom affiché de l'expéditeur
    """
    api_key = os.environ.get("SENDINBLUE_API_KEY", "")
    sender_email = os.environ.get("SENDINBLUE_SENDER_EMAIL", "")
    sender_name = os.environ.get("SENDINBLUE_SENDER_NAME", "")

    if not api_key:
        return "ERREUR : variable SENDINBLUE_API_KEY non configurée."
    if not sender_email:
        return "ERREUR : variable SENDINBLUE_SENDER_EMAIL non configurée."

    emails = json.loads(emails_json)
    if isinstance(emails, dict):
        if "erreur" in emails:
            return f"ERREUR lors de la lecture des emails : {emails['erreur']}"
        # Accepte le retour complet de lire_emails_depuis_excel_drive {fichier, emails}
        if "emails" in emails:
            emails = emails["emails"]
    if not emails:
        return "Aucune adresse email à contacter."

    headers = {
        "accept": "application/json",
        "api-key": api_key,
        "content-type": "application/json",
    }

    nb_ok, nb_err, erreurs = 0, 0, []

    for email in emails:
        payload = {
            "sender": {"name": sender_name, "email": sender_email},
            "to": [{"email": email}],
            "subject": sujet,
            "textContent": contenu,
        }
        try:
            resp = requests.post(BREVO_API_URL, json=payload, headers=headers, timeout=30)
            if resp.status_code in (200, 201):
                nb_ok += 1
            else:
                nb_err += 1
                erreurs.append(f"{email} → HTTP {resp.status_code}: {resp.text[:120]}")
        except requests.RequestException as exc:
            nb_err += 1
            erreurs.append(f"{email} → Exception: {exc}")

    rapport = (
        f"Diffusion terminée : {nb_ok} email(s) envoyé(s) avec succès, "
        f"{nb_err} échec(s)."
    )
    if erreurs:
        rapport += "\nDétail des échecs :\n" + "\n".join(erreurs)
    return rapport
